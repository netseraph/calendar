"""日历制作的参数"""

from openpyxl.styles import Alignment, Border, Font, Side, PatternFill

# 内容起始行
START_ROW = 10
# 内容起始列
START_COLUMN = 2
# 每行行高
ROW_HEIGHT = 20
ROW_HEIGHT_SEPARATOR = 5
COLUMN_WIDTH = 4

# 内容12个月按AREA_ROW行AREA_COLUMN列排布
AREA_ROW = 4
AREA_COLUMN = 12 // AREA_ROW


# 标题字体
TITLEFONT1 = Font(size=20)
TITLEFONT2 = Font(size=10)
# TITLEFONT3 = Font( size=10)
# 内容字体
DETAILSFONT07 = Font(size=7)
DETAILSFONT07G = Font(size=7, color="c0c0c0")
DETAILSFONT08 = Font(size=8)
DETAILSFONT08G = Font(size=8, color="c0c0c0")
DETAILSFONT09 = Font(size=9)
DETAILSFONT09G = Font(size=9, color="c0c0c0")
DETAILSFONT10 = Font(size=10)
DETAILSFONT10G = Font(size=10, color="c0c0c0")
DETAILSFONT11 = Font(size=11)
DETAILSFONT11G = Font(size=11, color="c0c0c0")
DETAILSFONT12 = Font(size=12)
DETAILSFONT12G = Font(size=12, color="c0c0c0")
# 文字对齐方式
# 水平居中,垂直居中
CC_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 水平居左,垂直居中
LC_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
# 水平居左,垂直顶头
LT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
# 水平居左,垂直顶头
RT_ALIGN = Alignment(horizontal="right", vertical="top", wrap_text=True)
# 边框样式
__side = Side(style="thin")
# 外侧边框
ALL_BORDER = Border(left=__side, right=__side, top=__side, bottom=__side)
# 左右上边框
LRT_BORDER = Border(left=__side, right=__side, top=__side)
# 左右下边框
LRB_BORDER = Border(left=__side, right=__side, bottom=__side)
# 左右边框
LR_BORDER = Border(left=__side, right=__side)
# 左边框
L_BORDER = Border(left=__side)
# 右边框
R_BORDER = Border(right=__side)
# 左下右边框
LB_BORDER = Border(left=__side, bottom=__side)
# 右下边框
RB_BORDER = Border(right=__side, bottom=__side)

# 填充样式
FILL1 = PatternFill(fgColor="3B789A", fill_type="solid")
FILL2 = PatternFill(fgColor="70AFCE", fill_type="solid")
FILL3 = PatternFill(fgColor="A5DEF1", fill_type="solid")
# FILL4 = PatternFill(fgColor="ffffff", fill_type="solid")
# FILL5 = PatternFill(fgColor="ffffff", fill_type="solid")
GRAYFILL = PatternFill(fgColor="F0F0F0", fill_type="solid")

STYLEDICT = {
    1: (
        (
            "一月",
            "二月",
            "三月",
            "四月",
            "五月",
            "六月",
            "七月",
            "八月",
            "九月",
            "十月",
            "十一月",
            "十二月",
        ),
        ("一", "二", "三", "四", "五", "六", "日"),
    ),
    2: (
        (
            "一",
            "二",
            "三",
            "四",
            "五",
            "六",
            "七",
            "八",
            "九",
            "十",
            "十一",
            "十二",
        ),
        ("一", "二", "三", "四", "五", "六", "日"),
    ),
    3: (
        (
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ),
        (
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ),
    ),
    4: (
        (
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ),
        ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"),
    ),
    5: (
        (
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ),
        ("M", "T", "W", "T", "F", "S", "S"),
    ),
}
