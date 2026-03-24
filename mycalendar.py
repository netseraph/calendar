"""年度日历制作"""

import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import config as CONF
from holidays import HOLIDAYS


def is_holiday(year: int, month: int, day: int) -> tuple:
    """通过日期是否在假期字典中,确认是否假期"""
    _date = f"{year}-{month:02}-{day:02}"
    if _date in HOLIDAYS:
        _r = HOLIDAYS[_date]
    else:
        _r = (-1, "")
    return _r


def calendar_year(year: int, style: int = 1, workpath: str = ""):
    """完整年度日历制作"""
    _monthlist = CONF.STYLEDICT[style][0]
    _weeklist = CONF.STYLEDICT[style][1]

    # 打开一个excel文件
    _wb = Workbook()
    _wb.properties.title = f"{year}年日历"
    _wb.properties.creator = "netseraph"
    _wb.properties.author = "netseraph"
    _wb.properties.company = "netseraph Studio"

    # 制作全年sheet
    _title = f"{year}年"

    # 设置当前sheet
    _ws = _wb.active

    # 设置sheet标题
    _ws.title = f"{year}年"

    # 纸张设置
    # 设置纸张方向为纵向
    _ws.page_setup.orientation = "portrait"
    # 设置纸张大小 B5
    _ws.page_setup.paperSize = 13
    # 设置页边距单位incn,0.5incn=1.27cm,0.25incn=0.635cm
    _ws.page_margins = PageMargins(
        left=0.5, right=0.25, top=0.25, bottom=0.25, header=0, footer=0
    )
    # 设置将整个工作表打印在一页
    _ws.sheet_properties.pageSetUpPr.fitToPage = True
    # 水平居中，垂直居中
    _ws.print_options.horizontalCentered = True
    _ws.print_options.verticalCentered = True

    # 设置工作区域行数
    _rowcount = CONF.AREA_ROW * 10 + 2
    # 设置工作区域列数
    _columncount = CONF.AREA_COLUMN * 8 + CONF.AREA_COLUMN - 1

    # 设置工作区域每列列宽
    for i in range(_columncount):
        _ws.column_dimensions[get_column_letter(i + 1)].width = CONF.COLUMN_WIDTH

    # 设置打印区域
    _top_left = f"{get_column_letter(1)}{2}"
    _bottom_right = f"{get_column_letter(_columncount)}{_rowcount+1}"
    _ws.print_area = f"{_top_left}:{_bottom_right}"

    # 设置视图模式
    _ws.sheet_view.view = "pageBreakPreview"
    # 隐藏网格线
    _ws.sheet_view.showGridLines = False
    # 显示比例
    _ws.sheet_view.zoomScale = 85

    # 标题-年
    # 设置标题行高
    _ws.row_dimensions[2].height = CONF.ROW_HEIGHT * 2
    # 合并设置标题行单元格
    _ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=_columncount,
    )
    # 输入标题行内容
    _ws.cell(row=2, column=1, value=f"{year}年")
    # 设置对齐方式
    _ws.cell(row=2, column=1).alignment = CONF.CC_ALIGN
    # 设置字体
    _ws.cell(row=2, column=1).font = CONF.TITLEFONT1

    # 设置标题分割行高
    _ws.row_dimensions[3].height = CONF.ROW_HEIGHT_SEPARATOR

    # 将12个月分为AERA_ROW行CONF.AREA_COLUMN列12个区域
    # 区域行循环
    for _row in range(CONF.AREA_ROW):
        # 当前起始行
        _cur_row = 4 + 10 * _row

        # 区域列循环
        for _column in range(CONF.AREA_COLUMN):
            # 当前起始列
            _cur_column = 1 + 9 * _column

            # 标题-月
            _cur_month = _row * CONF.AREA_COLUMN + _column
            # 月标题行高
            _ws.row_dimensions[_cur_row].height = CONF.ROW_HEIGHT
            # 合并月标题
            _ws.merge_cells(
                start_row=_cur_row,
                start_column=_cur_column + 1,
                end_row=_cur_row,
                end_column=_cur_column + 7,
            )
            # 月标题行内容
            _ws.cell(row=_cur_row, column=_cur_column + 1, value=_monthlist[_cur_month])
            # 对齐方式
            _ws.cell(row=_cur_row, column=_cur_column + 1).alignment = CONF.LC_ALIGN
            # 填充方式
            _ws.cell(row=_cur_row, column=_cur_column + 1).fill = CONF.FILL2
            # 字体
            _ws.cell(row=_cur_row, column=_cur_column + 1).font = CONF.TITLEFONT2

            # 间隔行高
            _ws.row_dimensions[_cur_row + 1].height = CONF.ROW_HEIGHT_SEPARATOR

            # 标题-周
            for i in range(7):
                _r = _cur_row + 2
                _c = _cur_column + i + 1
                # 周标题行高
                _ws.row_dimensions[_r].height = CONF.ROW_HEIGHT
                # 周标题行高内容
                _ws.cell(row=_r, column=_c, value=_weeklist[i])
                # 周标题对齐方式
                _ws.cell(row=_r, column=_c).alignment = CONF.CC_ALIGN
                # 周标题填充方式
                _ws.cell(row=_r, column=_c).fill = CONF.FILL3
                # 周标题字体
                _ws.cell(row=_r, column=_c).font = CONF.DETAILSFONT08

            # 内容-日期
            # 本月第一天
            _firstdayofmonth = datetime.date(year, _cur_month + 1, 1)
            # 每月第一天与周一之间的偏移
            _offset = datetime.date(year, _cur_month + 1, 1).weekday()
            # 计算本周第一天
            _cur_day = _firstdayofmonth - datetime.timedelta(days=_offset)

            # 循环填入日期
            for i in range(6):
                # 设置日期每行行高
                _r = _cur_row + i + 3
                _ws.row_dimensions[_r].height = CONF.ROW_HEIGHT

                # 本周是全年第几周
                # 内容
                _ws.cell(
                    row=_r, column=_cur_column, value=f"W{_cur_day.isocalendar()[1]}"
                )
                # 对齐方式
                _ws.cell(row=_r, column=_cur_column).alignment = CONF.CC_ALIGN

                # 填充方式
                # _ws.cell(row=_r, column=_cur_column).fill = CONF.FILL4
                # 字体
                _ws.cell(row=_r, column=_cur_column).font = CONF.DETAILSFONT07

                for j in range(7):
                    # 假期状态，0=假期，1=上班,-1=平常日期
                    _holiday = is_holiday(_cur_day.year, _cur_day.month, _cur_day.day)
                    _c = _cur_column + j + 1
                    # 日期单元格内容
                    _ws.cell(row=_r, column=_c, value=_cur_day.day)
                    # 日期单元格对齐方式
                    _ws.cell(row=_r, column=_c).alignment = CONF.CC_ALIGN
                    ## 日期单元格内容
                    # _ws.cell(row=_r, column=_c).fill = CONF.FILL4
                    # 如果当前日期是本月日期用一种字体,否则使用另一种字体
                    if _cur_day.month == _cur_month + 1:
                        _ws.cell(row=_r, column=_c).font = CONF.DETAILSFONT10
                    else:
                        _ws.cell(row=_r, column=_c).font = CONF.DETAILSFONT10G
                    # 根据是否休息,或者非上班的周末设置单元格填充

                    if _holiday[0] == 0 or (j in (5, 6) and _holiday[0] == -1):
                        _ws.cell(row=_r, column=_c).fill = CONF.GRAYFILL

                    _cur_day = _cur_day + datetime.timedelta(days=1)
        _ws.row_dimensions[_cur_row + 9].height = CONF.ROW_HEIGHT
        # print(_cur_row + 9)

    # 设置tips
    _ws.cell(row=1, column=1, value="打印前，请设置表格中文字的字体。")

    # 制作每月sheet
    for _month in range(1, 13):
        # 新建工作表
        _ws = _wb.create_sheet(title=f"{_month:02}月")

        # 设置页边距单位incn, 0.5incn=1.27cm,0.25incn=0.635cm
        _ws.page_margins = PageMargins(
            left=0.5, right=0.25, top=0.25, bottom=0.25, header=0, footer=0
        )

        # 纸张设置
        # 设置纸张方向为纵向
        _ws.page_setup.orientation = "portrait"
        # 设置纸张大小 B5
        _ws.page_setup.paperSize = 13
        # 设置页边距单位incn, 0.5incn=1.27cm,0.25incn=0.635cm
        _ws.page_margins = PageMargins(
            left=0.5, right=0.25, top=0.25, bottom=0.25, header=0, footer=0
        )
        # 设置将整个工作表打印在一页
        _ws.sheet_properties.pageSetUpPr.fitToPage = True
        # 水平居中，垂直居中
        _ws.print_options.horizontalCentered = True
        _ws.print_options.verticalCentered = True

        # 设置打印区域
        _top_left = f"{get_column_letter(1)}{2}"
        _bottom_right = f"{get_column_letter(14)}{15}"
        _ws.print_area = f"{_top_left}:{_bottom_right}"

        # 设置视图模式
        _ws.sheet_view.view = "pageBreakPreview"
        # 隐藏网格线
        _ws.sheet_view.showGridLines = False
        # 显示比例
        _ws.sheet_view.zoomScale = 85

        # 标题-年月
        _ws.row_dimensions[2].height = 40
        _ws.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=14,
        )
        _ws.cell(2, 1, value=f"{year}年{_month:02}月")
        _ws.cell(2, 1).alignment = CONF.CC_ALIGN
        _ws.cell(2, 1).font = CONF.TITLEFONT1

        # 标题-星期
        _ws.row_dimensions[3].height = 20
        for i in range(7):
            _columnletter = get_column_letter(i * 2 + 1)
            _ws.column_dimensions[_columnletter].width = 3
            _columnletter = get_column_letter(i * 2 + 2)
            _ws.column_dimensions[_columnletter].width = 9.5
            _ws.merge_cells(
                start_row=3,
                start_column=i * 2 + 1,
                end_row=3,
                end_column=i * 2 + 2,
            )
            _ws.cell(3, i * 2 + 1, value=_weeklist[i])
            _ws.cell(3, i * 2 + 1).alignment = CONF.CC_ALIGN
            _ws.cell(3, i * 2 + 1).font = CONF.TITLEFONT2
            _ws.cell(3, i * 2 + 1).border = CONF.ALL_BORDER
            _ws.cell(3, i * 2 + 2).border = CONF.ALL_BORDER

            if i in (5, 6):
                _ws.cell(3, i * 2 + 1).fill = CONF.GRAYFILL

        # 内容-日期
        _offset = datetime.date(year, _month, 1).weekday()
        _cur_day = datetime.date(year, _month, 1) - datetime.timedelta(days=_offset)

        for i in range(6):
            _ws.row_dimensions[i * 2 + 4].height = 20
            _ws.row_dimensions[i * 2 + 5].height = 100
            _cur_row = i * 2 + 4

            for j in range(7):
                _cur_column = j * 2 + 1
                # 日期单元格
                _ws.cell(_cur_row, _cur_column, value=_cur_day.day)
                # 设置文字对其方式
                _ws.cell(_cur_row, _cur_column).alignment = CONF.LT_ALIGN
                # 设置边框
                _ws.cell(_cur_row, _cur_column).border = CONF.L_BORDER

                # 假期单元格
                _holiday = is_holiday(_cur_day.year, _cur_day.month, _cur_day.day)
                _ws.cell(_cur_row, _cur_column + 1, value=_holiday[1])
                # 设置文字对其方式
                _ws.cell(_cur_row, _cur_column + 1).alignment = CONF.RT_ALIGN
                # 设置边框
                _ws.cell(_cur_row, _cur_column + 1).border = CONF.R_BORDER

                # 内容单元格
                _ws.cell(_cur_row + 1, _cur_column).border = CONF.LB_BORDER
                _ws.cell(_cur_row + 1, _cur_column + 1).border = CONF.RB_BORDER

                # 根据是否当前月设置文字格式
                if _cur_day.month == _month:
                    _ws.cell(_cur_row, _cur_column).font = CONF.DETAILSFONT08
                    _ws.cell(_cur_row, _cur_column + 1).font = CONF.DETAILSFONT08
                else:
                    _ws.cell(_cur_row, _cur_column).font = CONF.DETAILSFONT08G
                    _ws.cell(_cur_row, _cur_column + 1).font = CONF.DETAILSFONT08G

                # 根据是否休息,或者非上班的周末设置单元格填充
                if _holiday[0] == 0 or (j in (5, 6) and _holiday[0] == -1):
                    _ws.cell(_cur_row, _cur_column).fill = CONF.GRAYFILL
                    _ws.cell(_cur_row, _cur_column + 1).fill = CONF.GRAYFILL
                    _ws.cell(_cur_row + 1, _cur_column).fill = CONF.GRAYFILL
                    _ws.cell(_cur_row + 1, _cur_column + 1).fill = CONF.GRAYFILL

                _cur_day = _cur_day + datetime.timedelta(days=1)
        # 设置tips
        _ws.cell(row=1, column=1, value="打印前，请设置表格中文字的字体。")
    # 保存文件
    _wb.save(f"{workpath}/Calendar_{year}.xlsx")
