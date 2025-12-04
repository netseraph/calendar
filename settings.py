"""日历制作的常用参数"""

from openpyxl.styles import Alignment, Border, Font, Side, PatternFill

# 内容起始行
START_ROW = 10
# 内容列
START_COLUMN = 8
# 每行行高
ROW_HEIGHT = 20
ROW_HEIGHT_SEPARATOR = 5
COLUMN_WIDTH = 4
# 内容12个月按AREA_ROW行AREA_COLUMN列排布
AREA_ROW = 4
AREA_COLUMN = 12 // AREA_ROW

# 字体
# 标题字体
TITLEFONT1 = Font("Helvetica", size=20, bold=True)
TITLEFONT2 = Font("Helvetica", size=10, bold=True)
TITLEFONT3 = Font("Helvetica", size=10, bold=True)
# 内容字体
DETAILSFONT0 = Font("Helvetica", size=9)
DETAILSFONT1 = Font("Helvetica", size=9, color="f0f0f0")
# 文字对齐方式
# 水平居中,垂直居中
CCALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 水平居左,垂直居中
LCALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
# 边框样式
__side = Side(style="thin")
# 外侧边框
BORDER = Border(left=__side, right=__side, top=__side, bottom=__side)
# 填充样式
FILL1 = PatternFill(fgColor="3b789a", fill_type="solid")
FILL2 = PatternFill(fgColor="70afce", fill_type="solid")
FILL3 = PatternFill(fgColor="a5def1", fill_type="solid")
FILL4 = PatternFill(fgColor="ffffff", fill_type="solid")
FILL5 = PatternFill(fgColor="ffffff", fill_type="solid")

MONTHDICT = {
    1: (
        "一月",
        "二月",
        "三月",
        "四月",
        "五月",
        "六月",
        "七月",
        "八月",
        "九月",
        "十月",
        "十一月",
        "十二月",
    ),
    2: (
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ),
}

WEEKDICT = {
    1: ("一", "二", "三", "四", "五", "六", "日"),
    2: ("M", "T", "W", "T", "F", "S", "S"),
    3: ("Mon.", "Tues.", "Wed.", "Thur.", "Fri.", "Sat.", "Sun."),
    4: ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"),
}
