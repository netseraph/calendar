"""年度日历制作"""

import datetime

# import openpyxl

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import settings as SET


def calendar(year):
    """年度日历制作"""
    _monthlist = SET.MONTHDICT[2]
    _weeklist = SET.WEEKDICT[2]
    # 打开一个excel文件
    _wb = Workbook()
    # 设置当前sheet
    _ws = _wb.active

    # 设置
    _ws.title = f"{year}日历"

    # 纸张设置
    # 设置纸张方向纵向
    _ws.page_setup.orientation = _ws.ORIENTATION_PORTRAIT
    # 设置纸张大小B5
    _ws.page_setup.paperSize = 13

    # 设置将工作表调整为一页
    _ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 打印页面设置
    # 水平居中
    _ws.print_options.horizontalCentered = True
    # 垂直居中
    _ws.print_options.verticalCentered = True

    # 设置工作区域行数
    _rowcount = 2 + 10 * SET.AREA_ROW
    print(f"工作区共域{_rowcount}行")
    # 设置工作区域列数
    _columncount = 8 * SET.AREA_COLUMN + (SET.AREA_COLUMN - 1)
    print(f"工作区共域{_columncount}列")
    # 设置工作区域每列列宽
    for i in range(_columncount):
        _ws.column_dimensions[get_column_letter(SET.START_COLUMN + i)].width = (
            SET.COLUMN_WIDTH
        )

    # 设置打印区域
    _top_left = f"{get_column_letter(SET.START_COLUMN)}{SET.START_ROW}"
    _bottom_right = (
        f"{get_column_letter(SET.START_COLUMN+_columncount-1)}"
        f"{SET.START_ROW+_rowcount-1}"
    )
    _printarea = f"{_top_left}:{_bottom_right}"
    print(f"打印范围: {_printarea}")
    _ws.print_area = _printarea

    # 内容部分
    # 标题-年
    _ws.row_dimensions[SET.START_ROW].height = SET.ROW_HEIGHT * 2
    _ws.row_dimensions[SET.START_ROW + 1].height = SET.ROW_HEIGHT_SEPARATOR
    _ws.merge_cells(
        start_row=SET.START_ROW,
        start_column=SET.START_COLUMN,
        end_row=SET.START_ROW,
        end_column=SET.START_COLUMN + _columncount - 1,
    )
    _ws[_top_left].value = year
    _ws[_top_left].alignment = SET.CCALIGN
    # _ws[_top_left].fill = SET.FILL1
    _ws[_top_left].font = SET.TITLEFONT1

    # 将12个月分为AERA_ROW行AREA_COLUMN列12个区域
    # 区域行循环
    for _row in range(SET.AREA_ROW):
        # 当前起始行
        _cur_row = SET.START_ROW + 2 + 10 * _row
        # print(f"当前起始行: {_cur_row}")

        # 区域列循环
        for _column in range(SET.AREA_COLUMN):
            # 当前起始列
            _cur_column = SET.START_COLUMN + 9 * _column
            # print(f"当前起始列: {_cur_column}")

            # 标题-月
            _cur_month = _row * SET.AREA_COLUMN + _column
            _ws.row_dimensions[_cur_row].height = SET.ROW_HEIGHT
            _ws.row_dimensions[_cur_row + 1].height = SET.ROW_HEIGHT_SEPARATOR
            # print(_monthlist[_cur_month])
            _ws.merge_cells(
                start_row=_cur_row,
                start_column=_cur_column + 1,
                end_row=_cur_row,
                end_column=_cur_column + 7,
            )
            _ws.cell(row=_cur_row, column=_cur_column + 1).value = _monthlist[
                _cur_month
            ]
            _ws.cell(row=_cur_row, column=_cur_column + 1).alignment = SET.LCALIGN
            _ws.cell(row=_cur_row, column=_cur_column + 1).fill = SET.FILL2
            _ws.cell(row=_cur_row, column=_cur_column + 1).font = SET.TITLEFONT2

            # 标题-周
            for i in range(7):
                _ws.row_dimensions[_cur_row + 2].height = SET.ROW_HEIGHT
                _ws.cell(row=_cur_row + 2, column=_cur_column + i + 1).value = (
                    _weeklist[i]
                )
                _ws.cell(row=_cur_row + 2, column=_cur_column + i + 1).alignment = (
                    SET.CCALIGN
                )
                _ws.cell(row=_cur_row + 2, column=_cur_column + i + 1).fill = SET.FILL3
                _ws.cell(row=_cur_row + 2, column=_cur_column + i + 1).font = (
                    SET.TITLEFONT3
                )

            # 内容-日期
            # 本月第一天
            _firstdayofmonth = datetime.date(year, _cur_month + 1, 1)
            # print(_firstdayofmonth)
            # 每月第一天和周一的偏移
            _offset = datetime.date(year, _cur_month + 1, 1).weekday()
            # print(f"本月第一天{_firstdayofmonth}是周{_weeklist[_offset]}")
            # 计算本周第一天
            _cur_day = _firstdayofmonth - datetime.timedelta(days=_offset)
            # print(f"本周第一天是{_cur_day},本周是全年第{_cur_day.isocalendar()[1]}周")
            # _firstday = datetime.day(year, _month, 1)
            # 循环填入日期
            for i in range(6):
                # 设置日期每行行高
                #                print(_cur_row + i + 3)
                _ws.row_dimensions[_cur_row + i + 3].height = SET.ROW_HEIGHT

                # 本周是全年第几周
                _ws.cell(row=_cur_row + i + 3, column=_cur_column).value = (
                    _cur_day.isocalendar()[1]
                )
                _ws.cell(row=_cur_row + i + 3, column=_cur_column).alignment = (
                    SET.CCALIGN
                )
                _ws.cell(row=_cur_row + i + 3, column=_cur_column).fill = SET.FILL4
                _ws.cell(row=_cur_row + i + 3, column=_cur_column).font = (
                    SET.DETAILSFONT0
                )

                for j in range(7):
                    _ws.cell(row=_cur_row + i + 3, column=_cur_column + j + 1).value = (
                        _cur_day.day
                    )
                    _ws.cell(
                        row=_cur_row + i + 3, column=_cur_column + j + 1
                    ).alignment = SET.CCALIGN
                    _ws.cell(row=_cur_row + i + 3, column=_cur_column + j + 1).fill = (
                        SET.FILL4
                    )
                    # 如果当前日期是本月日期用一种字体,否则使用另一种字体
                    if _cur_day.month == _cur_month + 1:
                        _ws.cell(
                            row=_cur_row + i + 3, column=_cur_column + j + 1
                        ).font = SET.DETAILSFONT0
                    else:
                        _ws.cell(
                            row=_cur_row + i + 3, column=_cur_column + j + 1
                        ).font = SET.DETAILSFONT1

                    _cur_day = _cur_day + datetime.timedelta(days=1)
        _ws.row_dimensions[_cur_row + 9].height = SET.ROW_HEIGHT
        # print(_cur_row + 9)

    print("完成")
    _wb.save(f"/MyWorks/{year}日历.xlsx")


if __name__ == "__main__":
    calendar(2026)
